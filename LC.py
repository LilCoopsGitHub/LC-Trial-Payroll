#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul 15 14:49:06 2022

@author: Alexandra N. Walker
Please do not redistribute, copy, etc. without
permission from the owner. Only usage rights
have gone to Olivia from West Linn
Little Cooperstown
"""
import openpyxl
from openpyxl.styles import colors, Font, Color, NamedStyle, Border, Side, Alignment, Fill, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
import numpy as np
#%%
# Colors
header_color = 'BDD6EE'
gray_color = 'bfbfbf'

# Fonts
header_font = Font(name='Arial Black', size=11, color='980000',bold=True)
not_names_red_font = Font(name='Calibri',size=11, color='980000',bold=True)
names_font = Font(name='Calibri',size=11, color='980000')

# Border styles
border_style = Side(border_style='thick',color='000000')
border_double = Side(border_style='double',color='000000')
border_single  = Side(border_style='thin',color='000000')

# Fill Patterns
#  light blue
fill_pattern = PatternFill(patternType='solid',fgColor=header_color)
#  yellow
fill_pattern_yellow = PatternFill(patternType='solid',fgColor='ffff00')
#  gray
fill_pattern_gray = PatternFill(patternType='solid',fgColor=gray_color)
#%%
# Initializing
filename = input("Input the file name. Make sure it is a .csv format and located in your Downloads folder (ex. /Users/alex/Downloads/payroll_export.csv). Be sure to include the whole path. \n")
#%%
top_45 = 0 # Not sure what the top 45 is? Keeping this just in case. Set to 1 to fix
total_hours = 0

wb = Workbook()
ws = wb.active
ws.title = "LC Testing"
wb.save("LCTrial.xlsx")

wb = openpyxl.load_workbook("LCTrial.xlsx")
ws = wb["LC Testing"]

ws['A%s' % (top_45+1)] = 'Total tipshare'
ws['A%s' % (top_45+1)].font = not_names_red_font
ws['A%s' % (top_45+1)].alignment = Alignment(horizontal='right')

ws['A%s' % (top_45+2)] = 'Employee'
ws['A%s' % (top_45+2)].alignment = Alignment(horizontal='center')
ws['B%s' % (top_45+2)] = 'Position(s)'
ws['B%s' % (top_45+2)].alignment = Alignment(horizontal='center')
ws['C%s' % (top_45+2)] = 'Hours'
ws['C%s' % (top_45+2)].alignment = Alignment(horizontal='center')
ws['D%s' % (top_45+2)] = 'Share'
ws['D%s' % (top_45+2)].alignment = Alignment(horizontal='center')

for let in ['A','B','C', 'D']:
    ws['%s%s' % (let,top_45+2)].font = not_names_red_font
    ws.column_dimensions["%s" % (let)].width = 40
    ws['A%s' % (3)].border = Border(top=border_single, bottom=border_single,
                             right=border_single, left = border_single)
    
wb.save("LCTrial.xlsx")

hosts_expo = []
cooks = []
servers = []
bartenders = []
boh = []
mgr = []
other = []
all_employees = {}

# Fill yellow background
ws['B1'].fill = fill_pattern_yellow
ws['C1'].fill = fill_pattern_yellow

# Border around tip share
ws['A1'].border = Border(top=border_style, bottom=border_style, 
                         left=border_style, right=border_double)
ws['B1'].border = Border(top=border_style, bottom=border_style,
                         right=border_style)

f = open(filename,'r')
for i, line in enumerate(f.readlines()):
    #Replace all commas with ':)'
    if i!=0:
        #print("First")
        #print(i)
        #print(line.replace('"','').split(','))
        
        listy = line.replace('"','').split(',')
        print(listy)
        last_name = str(listy[0])
        first_name = str(listy[1][1:])
        position = str(listy[2])
        hours = float(listy[4])
        #print("Position is :",position.lower())
        if 'tip pool' in last_name.lower():
            #Set that one cell to the tip pool amount
            #Need to probably change this to the *total pay* column
            ws['B1'] = '$%s' % (hours)
            ws['B1'].font = not_names_red_font
            ws['B1'].alignment = Alignment(horizontal='right')
        if 'cook' in position.lower() or 'dish' in position.lower():
            cooks.append([last_name,first_name,hours,position])
        if 'host' in position.lower() or 'expo' in position.lower():
            hosts_expo.append([last_name,first_name,hours,position])
        if 'bar' in position.lower():
            bartenders.append([last_name,first_name,hours,position])
        if 'shift' in position.lower() or 'server' in position.lower():
            servers.append([last_name,first_name,hours,position])
            #print("Yo! A server")
            #print(position)
            #print(position.lower())
        if 'general' or 'mgr' in position.lower():
            mgr.append([last_name,first_name,hours,position])
        if position.lower()!=None and 'tip pool' not in last_name:
            #Need to check not already in dictionary
            #Try and see if it's a subset already
            all_employees[last_name] = [first_name,position,hours,position]

hosts_expo_starting_row = 4

#Want to make this a function, have it output last row
def put_in_employees(starting_row,employees,section_name):
    global total_hours
    '''
    Puts in all of the type of employees in one section
    Merges cell afterwards
    
    Parameters
    ----------
    starting_row : First row for first employee
    employees : List of employees (per type)
    section_name : Type of employees
    
    Returns
    -------
    final row: Row after final employee number
    '''
    for row, host_expo in enumerate(employees):
        #Need to check for duplicates
        #Want to print these out/inform the user first
        #If have multiple roles, add hours together, put in positions together
        #print("Row is ",row)
        #print("Starting row ",starting_row)
        ws['A%s' % (row+starting_row)] = host_expo[1] + ' ' + host_expo[0]
        ws['A%s' % (row+starting_row)].alignment = Alignment(horizontal='left')
        ws['A%s' % (row+starting_row)].font = names_font
        
        ws['B%s' % (row+starting_row)] = host_expo[3]
        ws['B%s' % (row+starting_row)].alignment = Alignment(horizontal='right')
        ws['B%s' % (row+starting_row)].font = names_font
        
        ws['C%s' % (row+starting_row)] = host_expo[2]
        ws['C%s' % (row+starting_row)].alignment = Alignment(horizontal='right')
        ws['C%s' % (row+starting_row)].font = names_font
    #print(np.shape(employees))
    last_row = np.shape(employees)[0]-1
    #print("Last row is ",last_row)
    wb.save("LCTrial.xlsx")
    
    # Sum up all hours
    sum_hours = 0
    for row_num in range(0,last_row+1):
        sum_hours += ws['C%s' % (row_num+starting_row)].value
    
    total_hours+=sum_hours
    ws['C%s' % (row+starting_row+1)] = sum_hours
    wb.save("LCTrial.xlsx")
    
    #print(sum_hours)
    
    #IF THE SECTION IS HOST/EXPO
    for row_num in range(0,last_row+1):
        if 'boh' in section_name.lower():
            # Need to calculate cell C2
            if row_num==0:
                try:
                    ws['C1'] = '$%s' % (float((ws['B1'].value).strip('$'))/total_hours) 
                except ZeroDivisionError:
                    ws['C1'] = 0
            
            #print(row_num+starting_row)
            #print(sum_hours)
            try:
                ws['D%s' % (row_num+starting_row)] = (float(ws['C%s' % (row+starting_row)].value))*(float((ws['C1'].value).strip('$')))
            except AttributeError:
                ws['D%s' % (row_num+starting_row)] = (float(ws['C%s' % (row+starting_row)].value))*(float((ws['C1'].value)))
           
            ws['D%s' % (row_num+starting_row)].alignment = Alignment(horizontal='right')
            ws['D%s' % (row_num+starting_row)].font = names_font
        else:
            #print(starting_row)
            #print(last_row+1)
            #print(row_num+starting_row)
            ws['D%s' % (row_num+starting_row)] = '$0'
            ws['D%s' % (row_num+starting_row)].alignment = Alignment(horizontal='right')
            ws['D%s' % (row_num+starting_row)].font = names_font
    wb.save("LCTrial.xlsx")
    
    # Sum up all shares
    sum_shares = 0
    for row_num in range(0,last_row+1):
        try:
            sum_shares += float((ws.cell(row=row_num+starting_row,column=4).value).strip('$'))
        except Exception:
            sum_shares += float((ws.cell(row=row_num+starting_row,column=4).value))
    ws['D%s' % (row+starting_row+1)] = '$%s' % (sum_shares)
    
    wb.save("LCTrial.xlsx")

    return(last_row+starting_row+1)

#next_host_expo_row = put_in_employees(hosts_expo_starting_row, hosts_expo)
# Merge Cells
# Will want to define another function
def type_of_section(starting_row,employees,section_name):
    #Section names: HOST/EXPO, BAR, BOH, SERVER
    
    #Put in all of the employees
    next_row = put_in_employees(starting_row,employees,section_name)
    
    #Create section
    ws.merge_cells('A%s:D%s' % (starting_row-1,starting_row-1))
    
    ws['A%s' % (starting_row-1)] = section_name
    
    #Merge and create bottom cells
    #print(top_45+next_row+1)
    #print("-----")
    #print(top_45+next_row)
    ws.merge_cells('A%s:B%s' % (top_45+next_row,top_45+next_row)) #want to be 12
    for let in ['A','B','C','D']:
        ws['%s%s' % (let,starting_row-1)].font = header_font
        ws['%s%s' % (let,starting_row-1)].fill = fill_pattern
        ws['%s%s' % (let,starting_row-1)].alignment = Alignment(horizontal='center')
        ws['%s%s' % (let,starting_row-1)].border = Border(top=border_single,bottom=border_single,
                                                          left=border_single,right=border_single)
        if let!='B':
            
            ws['%s%s' % (let,top_45+next_row)].font = not_names_red_font
            ws['%s%s' % (let,top_45+next_row)].fill = fill_pattern_gray
            ws['%s%s' % (let,top_45+next_row)].border = Border(top=border_double, bottom=border_double, 
                                     left=border_double, right=border_double)
            if let=='A':
                ws['%s%s' % (let,top_45+next_row)].alignment = Alignment(horizontal='center')
                ws['A%s' % (top_45+next_row)] = 'TOTAL %s HOURS:' % (section_name)
                ws['B%s' % (top_45+next_row)].border = Border(top=border_double, bottom=border_double, left=border_double, right=border_double)
            else:
                ws['%s%s' % (let,top_45+next_row)].alignment = Alignment(horizontal='right')
                if let=='C':
                    # Sum up all of the host/expo hours
                    sum_hours = 0
                    for m in range(starting_row,next_row):
                        sum_hours += float(ws.cell(row=m,column=3).value)
                    ws['C%s' % (top_45+next_row)] = sum_hours
                if let=='D':
                    sum_shares = 0
                    for m in range(starting_row,next_row):
                        try:
                            sum_shares += float((ws.cell(row=m,column=4).value).strip('$'))
                        except Exception:
                            sum_shares += float((ws.cell(row=m,column=4).value))
                    ws['D%s' % (top_45+next_row)] = sum_shares
                
    return(next_row)
            
                    
host_end = type_of_section(hosts_expo_starting_row,hosts_expo,"HOSTS/EXPO")
server_end = type_of_section(host_end+2,servers,"SERVERS")
bar_end = type_of_section(server_end+1,bartenders,"BAR")
boh_end = type_of_section(bar_end+1,cooks,"BOH")

#print('boh_end ',boh_end)
#Creating cell for total hours
# Merge last four
ws['A%s' % (boh_end+1)] = 'TOTAL LC HOURS:'
ws['D%s' % (boh_end+1)] = total_hours
ws.merge_cells('A%s:C%s' % (boh_end+1,boh_end+1))

for let in ['A','B','C']:
    ws['%s%s' % (let,boh_end+1)].alignment = Alignment(horizontal='right')
    ws['%s%s' % (let,boh_end+1)].border = Border(top=border_double, bottom=border_double, left=border_double, right=border_double)
    ws['%s%s' % (let,boh_end+1)].font = not_names_red_font
    ws['%s%s' % (let,boh_end+1)].fill = fill_pattern_gray

ws['D%s' % (boh_end+1)].alignment = Alignment(horizontal='center')
ws['D%s' % (boh_end+1)].border = Border(top=border_double, bottom=border_double, left=border_double, right=border_double)
ws['D%s' % (boh_end+1)].font = not_names_red_font
ws['D%s' % (boh_end+1)].fill = fill_pattern_gray
#%%
wb.save("LCTrial.xlsx")
f.close()